#!/usr/bin/env python3
"""Inspect an XLSX workbook and summarize workbook-native structure."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - import guard
    raise SystemExit(
        "Missing Python package 'openpyxl'. Install it with "
        "`python3 -m pip install openpyxl` and retry."
    ) from exc


def range_string(ws) -> str:
    return f"{ws.min_row}:{ws.max_row} x {ws.min_column}:{ws.max_column}"


def count_formulas(ws) -> int:
    count = 0
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                count += 1
    return count


def summarize_sheet(ws) -> dict[str, Any]:
    tables = []
    table_container = getattr(ws, "tables", {})
    if hasattr(table_container, "values"):
        for table in table_container.values():
            tables.append({"name": getattr(table, "name", None), "ref": getattr(table, "ref", None)})
    else:
        for name, ref in getattr(table_container, "items", lambda: [])():
            tables.append({"name": name, "ref": getattr(ref, "ref", None)})

    validations = 0
    dv = getattr(ws, "data_validations", None)
    if dv is not None:
        validations = len(getattr(dv, "dataValidation", []) or [])

    conditional = len(getattr(ws.conditional_formatting, "_cf_rules", {}) or {})

    return {
        "title": ws.title,
        "state": ws.sheet_state,
        "dimensions": ws.calculate_dimension(),
        "size_index": range_string(ws),
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "merged_ranges": len(ws.merged_cells.ranges),
        "freeze_panes": str(ws.freeze_panes) if ws.freeze_panes else None,
        "auto_filter": getattr(ws.auto_filter, "ref", None),
        "tables": tables,
        "formula_count": count_formulas(ws),
        "data_validation_rules": validations,
        "conditional_format_regions": conditional,
        "chart_count": len(getattr(ws, "_charts", [])),
        "image_count": len(getattr(ws, "_images", [])),
        "print_area": str(ws.print_area) if ws.print_area else None,
    }


def summarize_workbook(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=False)
    defined_names = []
    dn_dict = getattr(wb.defined_names, "dict", {})
    for name, obj in dn_dict.items():
        defined_names.append(
            {
                "name": name,
                "hidden": getattr(obj, "hidden", None),
                "value": getattr(obj, "value", None),
            }
        )

    return {
        "path": str(path.resolve()),
        "sheet_count": len(wb.sheetnames),
        "sheet_names": list(wb.sheetnames),
        "defined_names": defined_names,
        "external_link_count": len(getattr(wb, "_external_links", []) or []),
        "sheets": [summarize_sheet(wb[name]) for name in wb.sheetnames],
    }


def print_text(summary: dict[str, Any]) -> None:
    print(f"Workbook: {summary['path']}")
    print(f"Sheets ({summary['sheet_count']}): {', '.join(summary['sheet_names'])}")
    print(f"Defined names: {len(summary['defined_names'])}")
    print(f"External links: {summary['external_link_count']}")
    print()
    for sheet in summary["sheets"]:
        print(f"[{sheet['title']}] state={sheet['state']} range={sheet['dimensions']} formulas={sheet['formula_count']}")
        print(
            f"  merged={sheet['merged_ranges']} tables={len(sheet['tables'])} "
            f"validations={sheet['data_validation_rules']} conditional={sheet['conditional_format_regions']}"
        )
        print(
            f"  freeze_panes={sheet['freeze_panes']} auto_filter={sheet['auto_filter']} print_area={sheet['print_area']}"
        )
        if sheet["tables"]:
            for table in sheet["tables"]:
                print(f"  table: {table['name']} {table['ref']}")
        print()


def main() -> int:
    parser = argparse.ArgumentParser(description="Inspect an XLSX workbook.")
    parser.add_argument("workbook", type=Path, help="Path to .xlsx workbook")
    parser.add_argument("--json", action="store_true", help="Emit JSON summary")
    args = parser.parse_args()

    if not args.workbook.is_file():
        raise SystemExit(f"Workbook not found: {args.workbook}")

    summary = summarize_workbook(args.workbook)
    if args.json:
        print(json.dumps(summary, ensure_ascii=False, indent=2))
    else:
        print_text(summary)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
